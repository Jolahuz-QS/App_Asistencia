# app.py
# ------------------------------------------------------------
# Sistema de asistencia por reunión (Flask + SQLite + Excel template)
# ✅ Link único por reunión + QR
# ✅ Admin: crear / editar / eliminar reunión
# ✅ Acciones: abrir registros, eliminar registros (asistentes), Acta/Cierre, cerrar/reabrir registro
# ✅ Acta/Cierre: Desarrollo (opcional), Observaciones (opcional), Proyectó, Elaborado por, N compromisos
# ✅ Export Excel: llena tu plantilla Formato_Asistencia-Plantilla.xlsx respetando celdas combinadas
#    - +20 participantes: inserta filas
#    - +3 compromisos: inserta filas (sin error StyleProxy)
#    - Evita errores MergedCell escribiendo SIEMPRE en la top-left del merge
# ✅ Autenticación + Roles:
#    - Admin (role=admin): todas las funciones
#    - users_cits (role=staff): crear reunión, editar reunión, acta/cierre, exportar Excel y cerrar registro
#
# USUARIOS CREADOS AUTOMÁTICAMENTE (si no existen):
#   1) users_cits / Asistencia_cits   (role=staff)
#   2) Admin      / Admincits@2026    (role=admin)
#
# REQUISITOS:
#   pip install flask openpyxl qrcode pillow werkzeug
#
# EJECUTAR:
#   python app.py
#   http://127.0.0.1:5000/login
# ------------------------------------------------------------

from __future__ import annotations

import io
import os
import re
import json
import shutil
import secrets
import sqlite3
from copy import copy
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import qrcode
from flask import (
    Flask,
    abort,
    redirect,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
    g,
)

from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# -----------------------------
# Config
# -----------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "dev-secret-key")  # en producción: variable de entorno

DB = "asistencia.db"
PLANTILLA_XLSX = "Formato_Asistencia-Plantilla.xlsx"

TIPOS_DOCUMENTO = ["CC", "CE", "TI", "NIT", "Pasaporte"]
TIPOS_ASISTENCIA = ["Presencial", "Virtual"]

# Columnas A..V (22) para bloques inferiores
COL_A = 1
COL_B = 2
COL_C = 3
COL_D = 4
COL_N = 14
COL_O = 15
COL_P = 16
COL_R = 18
COL_S = 19
COL_V = 22

# Roles
ROLE_ADMIN = "admin"
ROLE_STAFF = "staff"  # users_cits


# -----------------------------
# DB helpers
# -----------------------------
def db_connect():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con


def init_db():
    with db_connect() as con:
        cur = con.cursor()

        # --- Users
        cur.execute(
            """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
        )

        # --- Meetings
        cur.execute(
            """
        CREATE TABLE IF NOT EXISTS meetings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            token TEXT NOT NULL UNIQUE,

            proceso TEXT NOT NULL,
            responsable TEXT NOT NULL,
            lugar TEXT NOT NULL,
            fecha TEXT NOT NULL,
            hora_inicio TEXT NOT NULL,
            convocados TEXT,
            asunto TEXT NOT NULL,

            is_closed INTEGER NOT NULL DEFAULT 0,

            desarrollo TEXT,
            observaciones TEXT,
            proyecto TEXT,
            elaborado_por TEXT,
            hora_fin TEXT,

            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
        )

        # --- Migración ligera: agregar hora_fin si no existe
        try:
            cur.execute("ALTER TABLE meetings ADD COLUMN hora_fin TEXT")
        except sqlite3.OperationalError:
            pass


        # --- Attendance
        cur.execute(
            """
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            meeting_id INTEGER NOT NULL,

            nombres TEXT NOT NULL,
            apellidos TEXT NOT NULL,

            tipo_documento TEXT NOT NULL,
            numero_documento TEXT NOT NULL,

            programa_dependencia TEXT NOT NULL,
            cargo TEXT NOT NULL,
            email TEXT NOT NULL,
            telefono TEXT NOT NULL,

            tipo_asistencia TEXT NOT NULL,
            created_at TEXT NOT NULL,

            UNIQUE(meeting_id, tipo_documento, numero_documento),
            FOREIGN KEY(meeting_id) REFERENCES meetings(id)
        )
        """
        )

        # --- Commitments
        cur.execute(
            """
        CREATE TABLE IF NOT EXISTS commitments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            meeting_id INTEGER NOT NULL,
            actividad TEXT NOT NULL,
            responsable TEXT NOT NULL,
            fecha TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(meeting_id) REFERENCES meetings(id)
        )
        """
        )

        con.commit()

    # Crear usuarios por defecto (idempotente)
    ensure_default_users()


def db_one(sql: str, params: Tuple[Any, ...] = ()) -> Optional[sqlite3.Row]:
    with db_connect() as con:
        cur = con.cursor()
        cur.execute(sql, params)
        return cur.fetchone()


def db_all(sql: str, params: Tuple[Any, ...] = ()) -> List[sqlite3.Row]:
    with db_connect() as con:
        cur = con.cursor()
        cur.execute(sql, params)
        return cur.fetchall()


def db_exec(sql: str, params: Tuple[Any, ...] = ()) -> None:
    with db_connect() as con:
        cur = con.cursor()
        cur.execute(sql, params)
        con.commit()


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_default_users() -> None:
    """
    Crea 2 usuarios si no existen:
      - users_cits / Asistencia_cits (role=staff)
      - Admin / Admincits@2026 (role=admin)
    """
    defaults = [
        ("users_cits", "Asistencia_cits", ROLE_STAFF),
        ("Admin", "Admincits@2026", ROLE_ADMIN),
    ]
    with db_connect() as con:
        cur = con.cursor()
        for username, password, role in defaults:
            cur.execute("SELECT id FROM users WHERE username=?", (username,))
            row = cur.fetchone()
            if row:
                continue
            cur.execute(
                """
                INSERT INTO users(username, password_hash, role, is_active, created_at, updated_at)
                VALUES (?,?,?,?,?,?)
                """,
                (
                    username,
                    generate_password_hash(password),
                    role,
                    1,
                    now_iso(),
                    now_iso(),
                ),
            )
        con.commit()


# -----------------------------
# Auth helpers
# -----------------------------
def current_user() -> Optional[sqlite3.Row]:
    uid = session.get("user_id")
    if not uid:
        return None
    u = db_one("SELECT * FROM users WHERE id=? AND is_active=1", (uid,))
    return u


def login_required(fn):
    def wrapper(*args, **kwargs):
        if not g.user:
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)

    wrapper.__name__ = fn.__name__
    return wrapper


def role_required(allowed_roles: List[str]):
    def decorator(fn):
        def wrapper(*args, **kwargs):
            if not g.user:
                return redirect(url_for("login", next=request.path))
            if g.user["role"] not in allowed_roles:
                return abort(403)
            return fn(*args, **kwargs)

        wrapper.__name__ = fn.__name__
        return wrapper

    return decorator


@app.before_request
def _ensure_db_and_user():
    init_db()
    g.user = current_user()


# -----------------------------
# Util
# -----------------------------
def safe_filename(name: str) -> str:
    name = name.strip()
    name = re.sub(r'[\\/*?:"<>|]+', "_", name)
    name = re.sub(r"\s+", " ", name)
    return name[:160] if len(name) > 160 else name


def norm(s: str) -> str:
    s = str(s or "").strip().lower()
    s = (
        s.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ü", "u")
    )
    s = " ".join(s.split())
    return s


# -----------------------------
# Excel helpers (robustos)
# -----------------------------
def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 60) -> None:
    """
    Copia de estilos robusta para filas insertadas, evitando:
      TypeError: unhashable type: 'StyleProxy'
    """
    max_col = min(max_col, ws.max_column)

    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass

    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)

        try:
            dst._style = copy(src._style)
        except Exception:
            try:
                dst.font = copy(src.font)
            except Exception:
                pass
            try:
                dst.fill = copy(src.fill)
            except Exception:
                pass
            try:
                dst.border = copy(src.border)
            except Exception:
                pass
            try:
                dst.alignment = copy(src.alignment)
            except Exception:
                pass
            try:
                dst.protection = copy(src.protection)
            except Exception:
                pass

        try:
            dst.number_format = src.number_format
        except Exception:
            pass

        dst.comment = None


def merge_safe(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    target = (r1, c1, r2, c2)
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if not (
            rng.max_row < target[0]
            or rng.min_row > target[2]
            or rng.max_col < target[1]
            or rng.min_col > target[3]
        ):
            to_unmerge.append(str(rng))
    for a in to_unmerge:
        try:
            ws.unmerge_cells(a)
        except Exception:
            pass
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def top_left_of_merge(ws, r: int, c: int) -> Tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return (rng.min_row, rng.min_col)
    return (r, c)


def set_value_safe(ws, r: int, c: int, value: Any) -> None:
    rr, cc = top_left_of_merge(ws, r, c)
    ws.cell(rr, cc).value = value


def set_alignment_safe(ws, r: int, c: int, alignment: Alignment) -> None:
    rr, cc = top_left_of_merge(ws, r, c)
    ws.cell(rr, cc).alignment = alignment


def find_cell_contains(ws, text: str) -> Optional[Tuple[int, int]]:
    t = norm(text)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and t in norm(cell.value):
                return (cell.row, cell.column)
    return None


def find_header_row_near(ws, start_row: int, must_have_terms: List[str], scan_down: int = 30) -> Optional[int]:
    terms = [norm(x) for x in must_have_terms]
    for r in range(start_row, start_row + scan_down + 1):
        found = set()
        for c in range(1, min(ws.max_column, 150) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                nv = norm(v)
                for t in terms:
                    if t in nv:
                        found.add(t)
        if all(t in found for t in terms):
            return r
    return None


def find_col_by_header_contains(ws, header_row: int, header_text: str) -> Optional[int]:
    ht = norm(header_text)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and ht in norm(v):
            return c
    return None


def replace_placeholders(ws, mapping: Dict[str, str]) -> None:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value in mapping:
                set_value_safe(ws, cell.row, cell.column, mapping[cell.value])


# -----------------------------
# Construcción de bloques inferiores (definitivo)
# -----------------------------
def build_blocks_after_participants(
    ws,
    after_participants_row: int,
    desarrollo: str,
    compromisos: List[Dict[str, str]],
    observaciones: str,
    proyecto: str,
    elaborado_por: str,
) -> None:
    # 1) Título Desarrollo
    r_des_title = after_participants_row + 1
    merge_safe(ws, r_des_title, COL_A, r_des_title, COL_V)
    set_value_safe(ws, r_des_title, COL_A, "DESARROLLO DE LA REUNIÓN")
    set_alignment_safe(ws, r_des_title, COL_A, Alignment(horizontal="center", vertical="center"))

    # 2) Bloque desarrollo (5 filas)
    r_des_start = r_des_title + 1
    r_des_end = r_des_start + 4
    merge_safe(ws, r_des_start, COL_A, r_des_end, COL_V)
    set_value_safe(ws, r_des_start, COL_A, desarrollo or "")
    set_alignment_safe(ws, r_des_start, COL_A, Alignment(wrap_text=True, vertical="top", horizontal="left"))

    # 3) Título Compromisos
    r_comp_title = r_des_end + 1
    merge_safe(ws, r_comp_title, COL_A, r_comp_title, COL_V)
    set_value_safe(ws, r_comp_title, COL_A, "COMPROMISOS")
    set_alignment_safe(ws, r_comp_title, COL_A, Alignment(horizontal="center", vertical="center"))

    # 4) Encabezados compromisos
    r_head = r_comp_title + 1
    ws.cell(r_head, COL_A).value = "ID"
    ws.cell(r_head, COL_A).alignment = Alignment(horizontal="center", vertical="center")

    merge_safe(ws, r_head, COL_B, r_head, COL_O)
    set_value_safe(ws, r_head, COL_B, "ACTIVIDAD")
    set_alignment_safe(ws, r_head, COL_B, Alignment(horizontal="center", vertical="center"))

    merge_safe(ws, r_head, COL_P, r_head, COL_R)
    set_value_safe(ws, r_head, COL_P, "RESPONSABLE")
    set_alignment_safe(ws, r_head, COL_P, Alignment(horizontal="center", vertical="center"))

    merge_safe(ws, r_head, COL_S, r_head, COL_V)
    set_value_safe(ws, r_head, COL_S, "FECHA")
    set_alignment_safe(ws, r_head, COL_S, Alignment(horizontal="center", vertical="center"))

    # 5) Filas compromisos
    base_rows = 3
    n = len(compromisos)
    need_rows = max(base_rows, n if n > 0 else base_rows)

    r_comp_start = r_head + 1

    if need_rows > base_rows:
        extra = need_rows - base_rows
        insert_at = r_comp_start + base_rows
        ws.insert_rows(insert_at, amount=extra)
        src_style_row = r_comp_start + base_rows - 1
        for rr in range(insert_at, insert_at + extra):
            copy_row_style(ws, src_style_row, rr, max_col=COL_V)

    for i in range(need_rows):
        rr = r_comp_start + i
        merge_safe(ws, rr, COL_B, rr, COL_O)
        merge_safe(ws, rr, COL_P, rr, COL_R)
        merge_safe(ws, rr, COL_S, rr, COL_V)

    for i in range(need_rows):
        rr = r_comp_start + i
        if i < n:
            ws.cell(rr, COL_A).value = i + 1
            ws.cell(rr, COL_A).alignment = Alignment(horizontal="center", vertical="center")

            set_value_safe(ws, rr, COL_B, compromisos[i]["actividad"])
            set_value_safe(ws, rr, COL_P, compromisos[i]["responsable"])
            set_value_safe(ws, rr, COL_S, compromisos[i]["fecha"])

            set_alignment_safe(ws, rr, COL_B, Alignment(wrap_text=True, vertical="center", horizontal="left"))
            set_alignment_safe(ws, rr, COL_P, Alignment(wrap_text=True, vertical="center", horizontal="left"))
            set_alignment_safe(ws, rr, COL_S, Alignment(wrap_text=True, vertical="center", horizontal="center"))
        else:
            ws.cell(rr, COL_A).value = ""
            set_value_safe(ws, rr, COL_B, "")
            set_value_safe(ws, rr, COL_P, "")
            set_value_safe(ws, rr, COL_S, "")

    r_comp_end = r_comp_start + need_rows - 1

    # 6) Observaciones
    r_obs_title = r_comp_end + 1
    merge_safe(ws, r_obs_title, COL_A, r_obs_title, COL_V)
    set_value_safe(ws, r_obs_title, COL_A, "OBSERVACIONES")
    set_alignment_safe(ws, r_obs_title, COL_A, Alignment(horizontal="center", vertical="center"))

    r_obs_start = r_obs_title + 1
    r_obs_end = r_obs_start + 2
    merge_safe(ws, r_obs_start, COL_A, r_obs_end, COL_V)
    set_value_safe(ws, r_obs_start, COL_A, observaciones or "")
    set_alignment_safe(ws, r_obs_start, COL_A, Alignment(wrap_text=True, vertical="top", horizontal="left"))

    # 7) Firmas
    r_sig = r_obs_end + 1

    merge_safe(ws, r_sig, COL_A, r_sig, COL_C)
    set_value_safe(ws, r_sig, COL_A, "Proyectó")
    set_alignment_safe(ws, r_sig, COL_A, Alignment(horizontal="center", vertical="center"))

    # D:N para dejar O libre
    merge_safe(ws, r_sig, COL_D, r_sig, COL_N)
    set_value_safe(ws, r_sig, COL_D, proyecto or "")
    set_alignment_safe(ws, r_sig, COL_D, Alignment(wrap_text=True, vertical="center", horizontal="left"))

    set_value_safe(ws, r_sig, COL_O, "Elaborado por")
    ws.cell(r_sig, COL_O).alignment = Alignment(horizontal="center", vertical="center")

    merge_safe(ws, r_sig, COL_P, r_sig, COL_V)
    set_value_safe(ws, r_sig, COL_P, elaborado_por or "")
    set_alignment_safe(ws, r_sig, COL_P, Alignment(wrap_text=True, vertical="center", horizontal="left"))


# -----------------------------
# Auth routes
# -----------------------------
@app.get("/login")
def login():
    if g.user:
        return redirect(url_for("admin"))
    nxt = request.args.get("next", "/admin")
    return render_template_string(LOGIN_TEMPLATE, next=nxt)


@app.post("/login")
def login_post():
    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()
    nxt = request.form.get("next") or "/admin"

    u = db_one("SELECT * FROM users WHERE username=? AND is_active=1", (username,))
    if not u or not check_password_hash(u["password_hash"], password):
        return render_template_string(LOGIN_TEMPLATE, next=nxt, error="Usuario o contraseña incorrectos.")

    session.clear()
    session["user_id"] = int(u["id"])
    return redirect(nxt)


@app.get("/logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))


# -----------------------------
# Routes (Public)
# -----------------------------
@app.get("/")
def home():
    return redirect(url_for("admin"))


@app.get("/m/<token>")
def meeting_page(token: str):
    m = db_one("SELECT * FROM meetings WHERE token=?", (token,))
    if not m:
        abort(404)
    closed = int(m["is_closed"]) == 1
    return render_template_string(
        PUBLIC_TEMPLATE,
        m=m,
        closed=closed,
        TIPOS_DOCUMENTO=TIPOS_DOCUMENTO,
        TIPOS_ASISTENCIA=TIPOS_ASISTENCIA,
    )


@app.post("/m/<token>")
def checkin(token: str):
    m = db_one("SELECT * FROM meetings WHERE token=?", (token,))
    if not m:
        abort(404)
    if int(m["is_closed"]) == 1:
        return "⚠️ El registro está cerrado."

    tipo_documento = request.form.get("tipo_documento", "").strip()
    tipo_asistencia = request.form.get("tipo_asistencia", "").strip()

    if tipo_documento not in TIPOS_DOCUMENTO:
        return "⚠️ Tipo de documento inválido."
    if tipo_asistencia not in TIPOS_ASISTENCIA:
        return "⚠️ Tipo de asistencia inválido."

    meeting_id = int(m["id"])

    data = {
        "nombres": request.form.get("nombres", "").strip(),
        "apellidos": request.form.get("apellidos", "").strip(),
        "tipo_documento": tipo_documento,
        "numero_documento": request.form.get("numero_documento", "").strip(),
        "programa_dependencia": request.form.get("programa_dependencia", "").strip(),
        "cargo": request.form.get("cargo", "").strip(),
        "email": request.form.get("email", "").strip(),
        "telefono": request.form.get("telefono", "").strip(),
        "tipo_asistencia": tipo_asistencia,
    }

    if not all(data.values()):
        return "⚠️ Debes diligenciar todos los campos."

    try:
        db_exec(
            """
            INSERT INTO attendance(
              meeting_id, nombres, apellidos,
              tipo_documento, numero_documento,
              programa_dependencia, cargo, email, telefono,
              tipo_asistencia, created_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                meeting_id,
                data["nombres"],
                data["apellidos"],
                data["tipo_documento"],
                data["numero_documento"],
                data["programa_dependencia"],
                data["cargo"],
                data["email"],
                data["telefono"],
                data["tipo_asistencia"],
                now_iso(),
            ),
        )
        return "✅ Registro exitoso."
    except sqlite3.IntegrityError:
        return "⚠️ Ya existe un registro para este documento en esta reunión."


@app.get("/m/<token>/qr.png")
def meeting_qr(token: str):
    m = db_one("SELECT * FROM meetings WHERE token=?", (token,))
    if not m:
        abort(404)

    meeting_url = url_for("meeting_page", token=token, _external=True)
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(meeting_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


# -----------------------------
# Routes (Admin)
# -----------------------------
@app.get("/admin")
@login_required
def admin():
    meetings = db_all("SELECT * FROM meetings ORDER BY fecha DESC, hora_inicio DESC, id DESC")
    return render_template_string(ADMIN_TEMPLATE, meetings=meetings, user=g.user)


@app.post("/admin/create")
@role_required([ROLE_ADMIN, ROLE_STAFF])
def create_meeting():
    token = secrets.token_urlsafe(16)
    db_exec(
        """
        INSERT INTO meetings(
          token, proceso, responsable, lugar, fecha, hora_inicio, convocados, asunto,
          is_closed, created_at, updated_at
        ) VALUES (?,?,?,?,?,?,?, ?, 0, ?, ?)
        """,
        (
            token,
            request.form.get("proceso", "").strip(),
            request.form.get("responsable", "").strip(),
            request.form.get("lugar", "").strip(),
            request.form.get("fecha", "").strip(),
            request.form.get("hora_inicio", "").strip(),
            (request.form.get("convocados", "").strip() or None),
            request.form.get("asunto", "").strip(),
            now_iso(),
            now_iso(),
        ),
    )
    return redirect(url_for("admin"))


@app.get("/admin/meeting/<int:meeting_id>/edit")
@role_required([ROLE_ADMIN, ROLE_STAFF])
def edit_meeting(meeting_id: int):
    m = db_one("SELECT * FROM meetings WHERE id=?", (meeting_id,))
    if not m:
        abort(404)
    return render_template_string(EDIT_TEMPLATE, m=m, user=g.user)


@app.post("/admin/meeting/<int:meeting_id>/edit")
@role_required([ROLE_ADMIN, ROLE_STAFF])
def save_edit_meeting(meeting_id: int):
    m = db_one("SELECT * FROM meetings WHERE id=?", (meeting_id,))
    if not m:
        abort(404)
    db_exec(
        """
        UPDATE meetings
        SET proceso=?, responsable=?, lugar=?, fecha=?, hora_inicio=?, convocados=?, asunto=?, updated_at=?
        WHERE id=?
        """,
        (
            request.form.get("proceso", "").strip(),
            request.form.get("responsable", "").strip(),
            request.form.get("lugar", "").strip(),
            request.form.get("fecha", "").strip(),
            request.form.get("hora_inicio", "").strip(),
            (request.form.get("convocados", "").strip() or None),
            request.form.get("asunto", "").strip(),
            now_iso(),
            meeting_id,
        ),
    )
    return redirect(url_for("admin"))


@app.post("/admin/meeting/<int:meeting_id>/delete")
@role_required([ROLE_ADMIN])  # SOLO Admin
def delete_meeting(meeting_id: int):
    db_exec("DELETE FROM attendance WHERE meeting_id=?", (meeting_id,))
    db_exec("DELETE FROM commitments WHERE meeting_id=?", (meeting_id,))
    db_exec("DELETE FROM meetings WHERE id=?", (meeting_id,))
    return redirect(url_for("admin"))


@app.post("/admin/meeting/<int:meeting_id>/delete-records")
@role_required([ROLE_ADMIN])  # SOLO Admin
def delete_records(meeting_id: int):
    db_exec("DELETE FROM attendance WHERE meeting_id=?", (meeting_id,))
    return redirect(url_for("admin"))


@app.get("/admin/meeting/<int:meeting_id>/records")
@role_required([ROLE_ADMIN])  # SOLO Admin (users_cits no lo pediste)
def records(meeting_id: int):
    m = db_one("SELECT * FROM meetings WHERE id=?", (meeting_id,))
    if not m:
        abort(404)
    rows = db_all(
        """
        SELECT * FROM attendance
        WHERE meeting_id=?
        ORDER BY created_at ASC
        """,
        (meeting_id,),
    )
    return render_template_string(RECORDS_TEMPLATE, m=m, rows=rows, user=g.user)


@app.get("/admin/meeting/<int:meeting_id>/minutes")
@role_required([ROLE_ADMIN, ROLE_STAFF])
def minutes(meeting_id: int):
    m = db_one("SELECT * FROM meetings WHERE id=?", (meeting_id,))
    if not m:
        abort(404)
    comps = db_all("SELECT * FROM commitments WHERE meeting_id=? ORDER BY id ASC", (meeting_id,))
    comps_payload = [{"actividad": c["actividad"], "responsable": c["responsable"], "fecha": c["fecha"]} for c in comps]
    return render_template_string(
        MINUTES_TEMPLATE,
        m=m,
        comps_json=json.dumps(comps_payload, ensure_ascii=False),
        user=g.user,
    )


@app.post("/admin/meeting/<int:meeting_id>/minutes/save")
@role_required([ROLE_ADMIN, ROLE_STAFF])
def minutes_save(meeting_id: int):
    m = db_one("SELECT * FROM meetings WHERE id=?", (meeting_id,))
    if not m:
        abort(404)

    desarrollo = request.form.get("desarrollo", "").strip() or None
    observaciones = request.form.get("observaciones", "").strip() or None
    proyecto = request.form.get("proyecto", "").strip() or None
    elaborado_por = request.form.get("elaborado_por", "").strip() or None
    hora_fin = request.form.get("hora_fin", "").strip() or None

    comps_raw = request.form.get("commitments_json", "[]").strip()
    try:
        comps = json.loads(comps_raw)
        if not isinstance(comps, list):
            comps = []
    except Exception:
        comps = []

    cleaned: List[Dict[str, str]] = []
    for x in comps:
        if not isinstance(x, dict):
            continue
        act = str(x.get("actividad", "")).strip()
        resp = str(x.get("responsable", "")).strip()
        fec = str(x.get("fecha", "")).strip()
        if act and resp and fec:
            cleaned.append({"actividad": act, "responsable": resp, "fecha": fec})

    db_exec(
        """
        UPDATE meetings
        SET desarrollo=?, observaciones=?, proyecto=?, elaborado_por=?, hora_fin=?, updated_at=?
        WHERE id=?
        """,
        (desarrollo, observaciones, proyecto, elaborado_por, hora_fin, now_iso(), meeting_id),
    )

    db_exec("DELETE FROM commitments WHERE meeting_id=?", (meeting_id,))
    for row in cleaned:
        db_exec(
            """
            INSERT INTO commitments(meeting_id, actividad, responsable, fecha, created_at)
            VALUES (?,?,?,?,?)
            """,
            (meeting_id, row["actividad"], row["responsable"], row["fecha"], now_iso()),
        )

    return redirect(url_for("minutes", meeting_id=meeting_id))


@app.post("/admin/meeting/<int:meeting_id>/close")
@role_required([ROLE_ADMIN, ROLE_STAFF])  # users_cits SÍ puede cerrar
def close_meeting(meeting_id: int):
    db_exec("UPDATE meetings SET is_closed=1, updated_at=? WHERE id=?", (now_iso(), meeting_id))
    return redirect(url_for("admin"))


@app.post("/admin/meeting/<int:meeting_id>/reopen")
@role_required([ROLE_ADMIN])  # SOLO Admin (según tu lista)
def reopen_meeting(meeting_id: int):
    db_exec("UPDATE meetings SET is_closed=0, updated_at=? WHERE id=?", (now_iso(), meeting_id))
    return redirect(url_for("admin"))


# -----------------------------
# Export
# -----------------------------
@app.get("/admin/meeting/<int:meeting_id>/export")
@role_required([ROLE_ADMIN, ROLE_STAFF])
def export_excel(meeting_id: int):
    m = db_one("SELECT * FROM meetings WHERE id=?", (meeting_id,))
    if not m:
        abort(404)

    if not os.path.exists(PLANTILLA_XLSX):
        return f"⚠️ No se encontró la plantilla '{PLANTILLA_XLSX}' junto a app.py"

    asistentes = db_all(
        """
        SELECT nombres, apellidos, tipo_documento, numero_documento,
               programa_dependencia, cargo, email, telefono, tipo_asistencia, created_at
        FROM attendance
        WHERE meeting_id=?
        ORDER BY created_at ASC
        """,
        (meeting_id,),
    )

    compromisos = db_all(
        """
        SELECT actividad, responsable, fecha
        FROM commitments
        WHERE meeting_id=?
        ORDER BY id ASC
        """,
        (meeting_id,),
    )

    os.makedirs("exports", exist_ok=True)
    out_name = safe_filename(f"Acta_{m['asunto']}_{m['fecha']}_{m['id']}.xlsx")
    out_path = os.path.join("exports", out_name)

    shutil.copyfile(PLANTILLA_XLSX, out_path)
    wb = load_workbook(out_path)
    ws = wb.active

    fixed_map = {
        "[PROCESO O DEPENDENCIA]": m["proceso"],
        "[RESPONSABLE]": m["responsable"],
        "[LUGAR DE LA REUNIÓN O LINK ]": m["lugar"],
        "[FECHA]": m["fecha"],
        "[HORA INICIO]": m["hora_inicio"],
        "[CONVOCADOS]": m["convocados"] or "",
        "[ASUNTO DE LA REUNIÓN]": m["asunto"],
    }
    replace_placeholders(ws, fixed_map)
    set_value_safe(ws, 7, 19, m["hora_fin"] or "")  # S7 Hora final

    pt = find_cell_contains(ws, "PARTICIPANTES")
    if not pt:
        return "⚠️ No encontré 'PARTICIPANTES' en la plantilla."
    pt_row = pt[0]

    header_row = find_header_row_near(ws, pt_row, ["NOMBRES", "APELLIDOS"], scan_down=50)
    if not header_row:
        return "⚠️ No encontré encabezados de participantes (NOMBRES/APELLIDOS)."

    data_start = header_row + 1

    des_pos = find_cell_contains(ws, "DESARROLLO DE LA REUN")
    if not des_pos:
        return "⚠️ No encontré el bloque 'DESARROLLO DE LA REUNIÓN' en la plantilla."
    des_title_row_template = des_pos[0]

    base_participant_rows = des_title_row_template - data_start
    if base_participant_rows <= 0:
        return "⚠️ La plantilla no tiene filas base para participantes."

    if len(asistentes) > base_participant_rows:
        extra = len(asistentes) - base_participant_rows
        insert_at = des_title_row_template
        ws.insert_rows(insert_at, amount=extra)

        src_style_row = data_start + base_participant_rows - 1
        for rr in range(insert_at, insert_at + extra):
            copy_row_style(ws, src_style_row, rr, max_col=COL_V)

        des_title_row_template += extra

    col_no = find_col_by_header_contains(ws, header_row, "No") or 1
    col_nombres = find_col_by_header_contains(ws, header_row, "NOMBRES")
    col_apellidos = find_col_by_header_contains(ws, header_row, "APELLIDOS")
    col_tipo_doc = find_col_by_header_contains(ws, header_row, "TIPO DE DOCUMENTO") or find_col_by_header_contains(ws, header_row, "TIPO")
    col_num_doc = find_col_by_header_contains(ws, header_row, "NÚMERO DE DOCUMENTO") or find_col_by_header_contains(ws, header_row, "NUMERO")
    col_prog = find_col_by_header_contains(ws, header_row, "PROGRAMA") or find_col_by_header_contains(ws, header_row, "DEPENDENCIA")
    col_cargo = find_col_by_header_contains(ws, header_row, "CARGO")
    col_email = find_col_by_header_contains(ws, header_row, "CORREO") or find_col_by_header_contains(ws, header_row, "ELECTRONICO")
    col_tel = find_col_by_header_contains(ws, header_row, "TEL") or find_col_by_header_contains(ws, header_row, "TELEFONO")
    col_tipo_asist = find_col_by_header_contains(ws, header_row, "TIPO ASISTENCIA")

    for i, a in enumerate(asistentes):
        r = data_start + i
        ws.cell(r, col_no).value = i + 1
        if col_nombres:
            ws.cell(r, col_nombres).value = a["nombres"]
        if col_apellidos:
            ws.cell(r, col_apellidos).value = a["apellidos"]
        if col_tipo_doc:
            ws.cell(r, col_tipo_doc).value = a["tipo_documento"]
        if col_num_doc:
            ws.cell(r, col_num_doc).value = a["numero_documento"]
        if col_prog:
            ws.cell(r, col_prog).value = a["programa_dependencia"]
        if col_cargo:
            ws.cell(r, col_cargo).value = a["cargo"]
        if col_email:
            ws.cell(r, col_email).value = a["email"]
        if col_tel:
            ws.cell(r, col_tel).value = a["telefono"]
        if col_tipo_asist:
            ws.cell(r, col_tipo_asist).value = a["tipo_asistencia"]

    last_participant_row = (data_start + len(asistentes) - 1) if len(asistentes) else (data_start - 1)
    gap_start = last_participant_row + 1
    gap_end = des_title_row_template - 1
    if gap_start <= gap_end:
        ws.delete_rows(gap_start, amount=(gap_end - gap_start + 1))
        des_title_row_template -= (gap_end - gap_start + 1)

    comps_list = [{"actividad": c["actividad"], "responsable": c["responsable"], "fecha": c["fecha"]} for c in compromisos]

    build_blocks_after_participants(
        ws=ws,
        after_participants_row=last_participant_row,
        desarrollo=(m["desarrollo"] or ""),
        compromisos=comps_list,
        observaciones=(m["observaciones"] or ""),
        proyecto=(m["proyecto"] or ""),
        elaborado_por=(m["elaborado_por"] or ""),
    )

    wb.save(out_path)
    return send_file(out_path, as_attachment=True)


# -----------------------------
# Templates
# -----------------------------
LOGIN_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Login</title>
  <style>
    body{font-family:Arial;max-width:520px;margin:60px auto;padding:0 16px}
    .card{border:1px solid #ddd;border-radius:12px;padding:16px}
    label{display:block;margin-top:10px}
    input,button{padding:10px;margin-top:6px;width:100%}
    button{cursor:pointer}
    .err{color:#b00020;margin-top:10px}
  </style>
</head>
<body>
  <h2>Iniciar sesión</h2>
  <div class="card">
    <form method="post" action="{{ url_for('login_post') }}">
      <input type="hidden" name="next" value="{{ next }}">
      <label>Usuario</label>
      <input name="username" autocomplete="username" required>
      <label>Contraseña</label>
      <input name="password" type="password" autocomplete="current-password" required>
      <button type="submit" style="margin-top:14px">Entrar</button>
      {% if error %}<div class="err">{{ error }}</div>{% endif %}
    </form>
  </div>
</body>
</html>
"""

PUBLIC_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Asistencia</title>
  <style>
    body{font-family:Arial;max-width:980px;margin:40px auto;padding:0 16px}
    .card{border:1px solid #ddd;border-radius:12px;padding:16px;margin-bottom:16px}
    .muted{color:#666;font-size:14px}
    label{display:block;margin-top:10px}
    input,select,button{width:100%;padding:10px;margin-top:6px}
    button{cursor:pointer}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
    @media (max-width: 860px){ .grid{grid-template-columns:1fr} }
    .badge{display:inline-block;padding:4px 10px;border-radius:999px;font-size:12px;border:1px solid #ddd}
  </style>
</head>
<body>
  <h2>Registro de asistencia</h2>

  <div class="card">
    <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start">
      <div>
        <div><b>{{ m['asunto'] }}</b></div>
        <div class="muted">
          {{ m['fecha'] }} {{ m['hora_inicio'] }} — {{ m['lugar'] }}<br>
          Proceso/Dependencia: {{ m['proceso'] }}<br>
          Responsable: {{ m['responsable'] }}
        </div>
      </div>
      {% if closed %}
        <div class="badge">Registro cerrado</div>
      {% else %}
        <div class="badge">Registro abierto</div>
      {% endif %}
    </div>
  </div>

  {% if closed %}
    <div class="card"><b>El registro está cerrado.</b></div>
  {% else %}
    <div class="card">
      <form method="post">
        <div class="grid">
          <div><label>Nombres</label><input name="nombres" required></div>
          <div><label>Apellidos</label><input name="apellidos" required></div>

          <div>
            <label>Tipo de documento</label>
            <select name="tipo_documento" required>
              <option value="" selected>Selecciona...</option>
              {% for td in TIPOS_DOCUMENTO %}<option value="{{ td }}">{{ td }}</option>{% endfor %}
            </select>
          </div>
          <div><label>Número de documento</label><input name="numero_documento" required></div>

          <div><label>Programa o dependencia</label><input name="programa_dependencia" required></div>
          <div><label>Cargo</label><input name="cargo" required></div>

          <div><label>Correo electrónico</label><input type="email" name="email" required></div>
          <div><label>Teléfono</label><input name="telefono" required></div>

          <div>
            <label>Tipo de asistencia</label>
            <select name="tipo_asistencia" required>
              <option value="" selected>Selecciona...</option>
              {% for ta in TIPOS_ASISTENCIA %}<option value="{{ ta }}">{{ ta }}</option>{% endfor %}
            </select>
          </div>
          <div></div>
        </div>
        <button type="submit" style="margin-top:14px">Registrar</button>
      </form>
    </div>
  {% endif %}
</body>
</html>
"""

ADMIN_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Admin</title>
  <style>
    body{font-family:Arial;max-width:1500px;margin:40px auto;padding:0 16px}
    .card{border:1px solid #ddd;border-radius:12px;padding:16px;margin-top:16px}
    label{display:block;margin-top:10px}
    input,button{padding:10px;margin-top:6px;width:100%}
    button{cursor:pointer}
    table{border-collapse:collapse;width:100%;margin-top:10px}
    th,td{border:1px solid #ddd;padding:10px;vertical-align:top}
    th{background:#f3f3f3}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
    @media (max-width: 1000px){ .grid{grid-template-columns:1fr} }
    img.qr{width:120px;height:120px;border:1px solid #eee;border-radius:8px}
    .badge{display:inline-block;padding:4px 10px;border-radius:999px;font-size:12px;border:1px solid #ddd}
    .btn{display:inline-block;padding:8px 10px;border:1px solid #ddd;border-radius:10px;text-decoration:none;color:#111;margin:2px 0}
    .btn-danger{border-color:#f0b4b4}
    .btn-ok{border-color:#bfe3bf}
    .btn-warn{border-color:#f0d69a}
    .mono{font-family: ui-monospace, Menlo, Consolas, monospace; font-size:12px}
    form.inline{display:inline}
    .topbar{display:flex;justify-content:space-between;align-items:center;gap:12px}
  </style>
</head>
<body>
  <div class="topbar">
    <h2 style="margin:0">Panel</h2>
    <div class="mono">
      Usuario: <b>{{ user['username'] }}</b> ({{ user['role'] }}) |
      <a href="{{ url_for('logout') }}">Salir</a>
    </div>
  </div>

  {% if user['role'] in ['admin','staff'] %}
  <div class="card">
    <h3 style="margin-top:0">Crear reunión</h3>
    <form method="post" action="{{ url_for('create_meeting') }}">
      <div class="grid">
        <div><label>Proceso o Dependencia</label><input name="proceso" required></div>
        <div><label>Responsable</label><input name="responsable" required></div>
        <div><label>Lugar de la reunión o link</label><input name="lugar" required></div>
        <div><label>Fecha</label><input type="date" name="fecha" required></div>
        <div><label>Hora de inicio</label><input type="time" name="hora_inicio" required></div>
        <div><label>Convocados (opcional)</label><input name="convocados"></div>
      </div>
      <label>Asunto de la reunión</label>
      <input name="asunto" required>
      <button type="submit" style="margin-top:14px">Crear</button>
    </form>
  </div>
  {% endif %}

  <div class="card">
    <h3 style="margin-top:0">Reuniones</h3>
    <table>
      <tr>
        <th>Estado</th>
        <th>Fecha/Hora</th>
        <th>Asunto</th>
        <th>Link/QR</th>
        <th>Acciones</th>
      </tr>
      {% for m in meetings %}
      <tr>
        <td>
          {% if m['is_closed']==1 %}
            <span class="badge">Cerrado</span>
          {% else %}
            <span class="badge">Abierto</span>
          {% endif %}
        </td>
        <td>{{ m['fecha'] }}<br>{{ m['hora_inicio'] }}</td>
        <td><b>{{ m['asunto'] }}</b><br><span class="mono">{{ m['proceso'] }}</span></td>
        <td>
          <div class="mono" style="margin-bottom:8px">{{ url_for('meeting_page', token=m['token'], _external=True) }}</div>
          <img class="qr" src="{{ url_for('meeting_qr', token=m['token']) }}">
        </td>
        <td style="white-space:nowrap">

          {% if user['role'] == 'admin' %}
            <a class="btn" href="{{ url_for('records', meeting_id=m['id']) }}">Abrir registros</a><br>
            <form class="inline" method="post" action="{{ url_for('delete_records', meeting_id=m['id']) }}"
                  onsubmit="return confirm('¿Eliminar TODOS los registros de asistencia de esta reunión?');">
              <button class="btn btn-danger" type="submit">Eliminar registros</button>
            </form><br>
          {% endif %}

          {% if user['role'] in ['admin','staff'] %}
            <a class="btn" href="{{ url_for('edit_meeting', meeting_id=m['id']) }}">Editar</a><br>
            <a class="btn btn-warn" href="{{ url_for('minutes', meeting_id=m['id']) }}">Acta/Cierre</a><br>
            <a class="btn btn-ok" href="{{ url_for('export_excel', meeting_id=m['id']) }}">Exportar Excel</a><br>
          {% endif %}

          {% if m['is_closed']==0 %}
            {% if user['role'] in ['admin','staff'] %}
            <form class="inline" method="post" action="{{ url_for('close_meeting', meeting_id=m['id']) }}">
              <button class="btn btn-warn" type="submit">Cerrar registro</button>
            </form><br>
            {% endif %}
          {% else %}
            {% if user['role'] == 'admin' %}
            <form class="inline" method="post" action="{{ url_for('reopen_meeting', meeting_id=m['id']) }}">
              <button class="btn btn-ok" type="submit">Reabrir registro</button>
            </form><br>
            {% endif %}
          {% endif %}

          {% if user['role'] == 'admin' %}
          <form class="inline" method="post" action="{{ url_for('delete_meeting', meeting_id=m['id']) }}"
                onsubmit="return confirm('¿Eliminar esta reunión y TODOS sus datos?');">
            <button class="btn btn-danger" type="submit">Eliminar reunión</button>
          </form>
          {% endif %}

        </td>
      </tr>
      {% endfor %}
    </table>
  </div>
</body>
</html>
"""

EDIT_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Editar reunión</title>
  <style>
    body{font-family:Arial;max-width:980px;margin:40px auto;padding:0 16px}
    .card{border:1px solid #ddd;border-radius:12px;padding:16px}
    label{display:block;margin-top:10px}
    input,button{padding:10px;margin-top:6px;width:100%}
    button{cursor:pointer}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
    @media (max-width: 860px){ .grid{grid-template-columns:1fr} }
    a{color:#111}
    .topbar{display:flex;justify-content:space-between;align-items:center;gap:12px}
  </style>
</head>
<body>
  <div class="topbar">
    <h2 style="margin:0">Editar reunión</h2>
    <div>
      <a href="{{ url_for('admin') }}">Volver</a> |
      <a href="{{ url_for('logout') }}">Salir</a>
    </div>
  </div>
  <div class="card">
    <form method="post">
      <div class="grid">
        <div><label>Proceso o Dependencia</label><input name="proceso" value="{{ m['proceso'] }}" required></div>
        <div><label>Responsable</label><input name="responsable" value="{{ m['responsable'] }}" required></div>
        <div><label>Lugar de la reunión o link</label><input name="lugar" value="{{ m['lugar'] }}" required></div>
        <div><label>Fecha</label><input type="date" name="fecha" value="{{ m['fecha'] }}" required></div>
        <div><label>Hora de inicio</label><input type="time" name="hora_inicio" value="{{ m['hora_inicio'] }}" required></div>
        <div><label>Convocados (opcional)</label><input name="convocados" value="{{ m['convocados'] or '' }}"></div>
      </div>
      <label>Asunto</label>
      <input name="asunto" value="{{ m['asunto'] }}" required>
      <button type="submit" style="margin-top:14px">Guardar</button>
    </form>
  </div>
</body>
</html>
"""

RECORDS_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Registros</title>
  <style>
    body{font-family:Arial;max-width:1500px;margin:40px auto;padding:0 16px}
    table{border-collapse:collapse;width:100%}
    th,td{border:1px solid #ddd;padding:10px;vertical-align:top}
    th{background:#f3f3f3}
    .mono{font-family: ui-monospace, Menlo, Consolas, monospace; font-size:12px}
  </style>
</head>
<body>
  <h2>Registros - {{ m['asunto'] }}</h2>
  <p class="mono">{{ url_for('meeting_page', token=m['token'], _external=True) }}</p>

  <table>
    <tr>
      <th>#</th><th>Nombres</th><th>Apellidos</th><th>Tipo Doc</th><th>Número</th>
      <th>Programa/Dependencia</th><th>Cargo</th><th>Email</th><th>Teléfono</th><th>Tipo asistencia</th><th>Fecha registro</th>
    </tr>
    {% for r in rows %}
    <tr>
      <td>{{ loop.index }}</td>
      <td>{{ r['nombres'] }}</td>
      <td>{{ r['apellidos'] }}</td>
      <td>{{ r['tipo_documento'] }}</td>
      <td>{{ r['numero_documento'] }}</td>
      <td>{{ r['programa_dependencia'] }}</td>
      <td>{{ r['cargo'] }}</td>
      <td>{{ r['email'] }}</td>
      <td>{{ r['telefono'] }}</td>
      <td>{{ r['tipo_asistencia'] }}</td>
      <td class="mono">{{ r['created_at'] }}</td>
    </tr>
    {% endfor %}
  </table>

  <p style="margin-top:12px"><a href="{{ url_for('admin') }}">← Volver</a></p>
</body>
</html>
"""

MINUTES_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Acta/Cierre</title>
  <style>
    body{font-family:Arial;max-width:1100px;margin:40px auto;padding:0 16px}
    .card{border:1px solid #ddd;border-radius:12px;padding:16px;margin-bottom:16px}
    label{display:block;margin-top:10px}
    input,textarea,button{padding:10px;margin-top:6px;width:100%}
    textarea{min-height:110px}
    button{cursor:pointer}
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
    @media (max-width: 900px){ .grid{grid-template-columns:1fr} }
    table{border-collapse:collapse;width:100%;margin-top:10px}
    th,td{border:1px solid #ddd;padding:10px;vertical-align:top}
    th{background:#f3f3f3}
    .row{display:flex;gap:10px;flex-wrap:wrap}
    .row .btn{width:auto}
    .btn{display:inline-block;padding:10px 12px;border:1px solid #ddd;border-radius:10px;background:#fff}
    .btn-danger{border-color:#f0b4b4}
    .btn-ok{border-color:#bfe3bf}
    .btn-warn{border-color:#f0d69a}
    .muted{color:#666;font-size:14px}
    .topbar{display:flex;justify-content:space-between;align-items:center;gap:12px}
    .mono{font-family: ui-monospace, Menlo, Consolas, monospace; font-size:12px}
  </style>
</head>
<body>
  <div class="topbar">
    <div>
      <h2 style="margin:0">Acta/Cierre - {{ m['asunto'] }}</h2>
      <div class="muted">{{ m['fecha'] }} {{ m['hora_inicio'] }} — {{ m['lugar'] }}</div>
    </div>
    <div class="mono">
      {{ user['username'] }} ({{ user['role'] }}) |
      <a href="{{ url_for('admin') }}">Panel</a> |
      <a href="{{ url_for('logout') }}">Salir</a>
    </div>
  </div>

  <div class="card">
    <form method="post" action="{{ url_for('minutes_save', meeting_id=m['id']) }}" onsubmit="syncCommitments()">
      <label>Desarrollo de la reunión (opcional)</label>
      <textarea name="desarrollo" id="desarrollo">{{ m['desarrollo'] or '' }}</textarea>

      <h3 style="margin-top:18px;margin-bottom:0">Compromisos</h3>
      <div class="muted">Agregar compromisos NO borra Observaciones. Solo se guarda al final.</div>

      <div class="row" style="margin-top:10px">
        <button type="button" class="btn btn-ok" onclick="addRow()">+ Agregar compromiso</button>
      </div>

      <table id="tbl">
        <thead>
          <tr>
            <th style="width:55%">Actividad</th>
            <th style="width:25%">Responsable</th>
            <th style="width:15%">Fecha</th>
            <th style="width:5%">Acción</th>
          </tr>
        </thead>
        <tbody></tbody>
      </table>

      <input type="hidden" name="commitments_json" id="commitments_json">

      <label style="margin-top:16px">Observaciones (opcional)</label>
      <textarea name="observaciones" id="observaciones">{{ m['observaciones'] or '' }}</textarea>

      <div class="grid">
      <div class="grid">
        <div>
          <label>Hora final</label>
          <input type="time" name="hora_fin" value="{{ m['hora_fin'] or '' }}">
        </div>
        <div></div>
      </div>

        <div><label>Proyectó</label><input name="proyecto" value="{{ m['proyecto'] or '' }}"></div>
        <div><label>Elaborado por</label><input name="elaborado_por" value="{{ m['elaborado_por'] or '' }}"></div>
      </div>

      <button type="submit" style="margin-top:14px">Guardar acta</button>
    </form>

    <div class="row" style="margin-top:12px">
      {% if m['is_closed']==0 %}
      <form method="post" action="{{ url_for('close_meeting', meeting_id=m['id']) }}" style="margin:0">
        <button class="btn btn-warn" type="submit">Cerrar registro</button>
      </form>
      {% else %}
      {% if user['role']=='admin' %}
      <form method="post" action="{{ url_for('reopen_meeting', meeting_id=m['id']) }}" style="margin:0">
        <button class="btn btn-ok" type="submit">Reabrir registro</button>
      </form>
      {% endif %}
      {% endif %}

      <a class="btn" href="{{ url_for('export_excel', meeting_id=m['id']) }}">Exportar Excel</a>

      {% if user['role']=='admin' %}
      <a class="btn" href="{{ url_for('records', meeting_id=m['id']) }}">Abrir registros</a>
      {% endif %}
    </div>
  </div>

<script>
  let data = {{ comps_json | safe }};

  function escapeHtml(s){
    return String(s||"").replace(/[&<>"']/g, (m)=>({ "&":"&amp;","<":"&lt;",">":"&gt;","\\"":"&quot;","'":"&#039;" }[m]));
  }

  function render(){
    const tb = document.querySelector("#tbl tbody");
    tb.innerHTML = "";
    data.forEach((r, idx)=>{
      const tr = document.createElement("tr");
      tr.innerHTML = `
        <td><input data-k="actividad" data-i="${idx}" value="${escapeHtml(r.actividad||"")}" /></td>
        <td><input data-k="responsable" data-i="${idx}" value="${escapeHtml(r.responsable||"")}" /></td>
        <td><input type="date" data-k="fecha" data-i="${idx}" value="${escapeHtml(r.fecha||"")}" /></td>
        <td><button type="button" class="btn btn-danger" onclick="delRow(${idx})">X</button></td>
      `;
      tb.appendChild(tr);
    });

    tb.querySelectorAll("input").forEach(inp=>{
      inp.addEventListener("input", (e)=>{
        const i = parseInt(e.target.getAttribute("data-i"));
        const k = e.target.getAttribute("data-k");
        data[i][k] = e.target.value;
      });
    });
  }

  function addRow(){
    data.push({actividad:"", responsable:"", fecha:""});
    render();
  }

  function delRow(i){
    data.splice(i, 1);
    render();
  }

  function syncCommitments(){
    document.getElementById("commitments_json").value = JSON.stringify(data);
  }

  render();
</script>
</body>
</html>
"""

# -----------------------------
# Main
# -----------------------------
if __name__ == "__main__":
    app.run(debug=True)